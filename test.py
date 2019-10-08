import initializer
import producer
import loss
import numpy as np

a = np.random.randn(10).__abs__()
types = ['ConcentrationLinear', 'Constant', 'ConcentrationSigmoid', 'WeightedDegreeLinear', 'WeightedDegreeSigmoid']

import producer
import BrainsphereModel

adjacency = np.loadtxt('/home/philipp/alzheimers/combined-atlas/euclidean_adjecency_6neighbors.txt')

connectivity = np.loadtxt(
    '/home/philipp/alzheimers/neurodegeneration-forecast/Network_Data/Control/Average/adj_threshold_6')

from openpyxl import load_workbook

wb = load_workbook(
    '/home/philipp/alzheimers/neurodegeneration-forecast/Neurodeg_Data_SUVR/SUVRs_New_Atlas_Modeling.xlsx')
ws = wb['SUVRs_New_Atlas']
concentrations = np.array([[i.value for i in j] for j in ws['B2':'UZ21']])
concentrations = concentrations[np.argsort(np.sum(concentrations, axis=1))]

# params = {"lowerInit": 0.3,
#          "upperInit": 0.6}  # 2149.3

# params = {'ConcentrationLinear': 2.6920626035910398, 'ProductionConstant': 0.7433074157266951, 'ConcentrationSigmoidA': 3.166102288632473, 'ConcentrationSigmoidB': 0.6100809472161937, 'ConcentrationSigmoidC': 7.449624824119648, 'WD-LinearA': 0.35120110860618303, 'WD-SigmoidA': 4.39411213859124, 'WD-SigmoidB': 7.652702177481957, 'WD-SigmoidC': 1.5091546379416079, 'lowerInit': 1.005106421821866, 'upperInit': 1.2086002841534664} #2052.006795851523
#params = {'lowerInit': 0.8, 'upperInit': 1.0}
params = {'ConcentrationLinear': 0.07, 'ProductionConstant': 0.01, 'ConcentrationSigmoidA': 1.0, 'ConcentrationSigmoidB': 1.0, 'ConcentrationSigmoidC': 2, 'WD-LinearA': 0.05, 'WD-SigmoidA': 0.1, 'WD-SigmoidB': 0.8194874373805783, 'WD-SigmoidC': 5, 'lowerInit': 0.9, 'upperInit': 1.20}


bsm = BrainsphereModel.BrainsphereModel(connectivity, concentrations, params=params, euclideanAdjacency=adjacency)

import Optimizer

o = Optimizer.Optimizer(bsm)
o.optimize()

# import timeit

# print(timeit.timeit('bsm.gradient()',number=10,globals=globals()))  # 96.5
# print(timeit.timeit('bsm.gradient4()',number=10,globals=globals())) # 99.0
# print(bsm.gradient4())
